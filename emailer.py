from pickle import FALSE
import openpyxl
import smtplib
from email.message import EmailMessage
import time

start = time.time()
#file Locs
#place destination of files here. Use \\ between folders.
emailList = "Insert path to email list here"
stockList = "insert path to stock list here"

#open workbooks
wbEmails = openpyxl.load_workbook(emailList)
print("email list loaded")
wbEmailsSheet = wbEmails.active
print("Stock list loading, please do not close. it is not frozen.")
wbStock = openpyxl.load_workbook(stockList)
print("stock list loaded")

#define area of emails list
emailsCells = wbEmailsSheet['A1':'B132']

server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
server.login("Insert your email", "Insert password")

#loop everything in here?
for row in emailsCells:
    SelectedEmailName = wbEmailsSheet.cell(row=row[0].row, column=1).value
    SelectedEmail = wbEmailsSheet.cell(row=row[0].row, column=2).value

    #select stock sheet from wbemails
    if SelectedEmailName in wbStock.sheetnames:
        wbStockSheet = wbStock.get_sheet_by_name(SelectedEmailName)

        #define area of stock list
        stockCells = wbStockSheet['A1':'F250']

        #printing values of stock sheet
        message = ""
        skip = False
        for cell1, cell2, cell3, cell4, cell5, cell6 in stockCells:
            if cell1.value == None and cell2.value == None and cell3.value == None and cell4.value == None and cell5.value == None and cell6.value == None and skip == True:
                break
            elif cell1.value == None and cell2.value == None and cell3.value == None and cell4.value == None and cell5.value == None and cell6.value == None and skip == False:
                skip = True

            if (cell1.value != None):
                one = cell1.value
                one = str(one)
                if one[0] in "0123456789":
                    one = (one[:10])
                if isinstance(one, str):
                    one = one.ljust(19,".")
            else:
                one = "..................."
            if (cell2.value != None):
                two = cell2.value
                two = two.ljust(7, ".")
                two = two + " |"
            else:
                two = "......."
            if (cell3.value != None):
                three = cell3.value
                three = three.ljust(11,".")
                three = three + " |"
            else:
                three = "..........."
            if (cell4.value != None):
                four = cell4.value
                four = four.ljust(46, ".")
            else:
                four = "."*46
            if (cell5.value != None):
                five = cell5.value
                five = str(five)
                five = five.ljust(14,".")
            else:
                five = "................"
            if (cell6.value != None):
                six = cell6.value
                six = str(six)
            else:
                six = "     "
            #print(one, two, three, four, five, six)
            message = message + "\n" + one +" "+ two +" "+ three +" "+ four +" "+ five +" "+ six +" "+ "\n" 
        print (message)
        message = message + "\n" + "This email was sent automatically. If there is an issue please reach out to davidhanys@gmail.com or alexanderhanys@gmail.com"
        msg = EmailMessage()
        msg.set_content(message)
        
        sendToEmail = wbEmailsSheet.cell(row=row[0].row, column=2).value
        msg['Subject'] = "ETA Report - " + SelectedEmailName
        msg['From'] = "Your Email here"
        msg['To'] = SelectedEmail

        # Send the message via SMTP server.
        server.send_message(msg)
    else:
        print("No sheet found for " + SelectedEmailName)
server.quit()
end = time.time()
print("Time taken: " + str(end - start))